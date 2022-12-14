import openpyxl


class HomePageData:
    test_Homepage_Data = {"first_name": "Test", "email": "Test@123.com", "gender": "Female"}, {
        "first_name": "SampleName", "email": "sampleName@123.com", "gender": "Male"}

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("\\PythonDemo.xlsx")
        Dict = {}
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    print(Dict)
